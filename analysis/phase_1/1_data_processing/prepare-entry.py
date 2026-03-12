#!/usr/bin/env python3

"""
prepare-entry.py
Author: Jared Johnson, jared.johnson@doh.wa.gov

Process genomic assembly metadata and FASTA files for submission preparation.
Reads Excel metadata, validates FASTA files, and exports structured data.
"""

import argparse
import csv
import gzip
import json
import logging
import os
import re
import shutil
import sys
import textwrap
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

VERSION = "1.1"

REQUIRED_COLUMNS = [
    'SRA ID',
    'Assembly Species / Segment',
    'Assembly Filename',
    'Assembly Generated?',
    'Sample ID',
    'Compute Environment (HPC cluster, cloud/local workstation)',
    'Workflow Name & Version',
    'Workflow Github Link (if hosted/public on GH)',
]

ASSEMBLY_GENERATED_VALUES = {'true', 'yes', '1', 'y'}


def sanitize_string(name: str, replacement: str = "_") -> str:
    if not name:
        return ""
    sanitized = re.sub(r'[<>:"/\\|?*\s]+', replacement, str(name))
    return sanitized.strip(replacement)[:255]


def ensure_outdir(outdir: Path) -> None:
    try:
        outdir.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        sys.exit(f"Error: Could not create output directory {outdir}: {e}")


def load_workbook_data(file_path: str) -> Tuple[str, Dict]:
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        sys.exit(f"Error: Could not load workbook {file_path}: {e}")

    sheet = 'Sheet1' if 'Sheet1' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    data = [
        [cell.strip() if isinstance(cell, str) else cell for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]

    if len(data) < 6:
        sys.exit("Error: Workbook does not have enough rows.")

    if not (data[3] and 'Submitter Name' in str(data[3][1] or '')):
        sys.exit('Error: Expected "Submitter Name" in row 4.')
    if not (data[5] and 'Sample ID' in str(data[5][1] or '')):
        sys.exit('Error: Expected "Sample ID" in row 6.')

    source = data[4][1] if data[4] and len(data[4]) > 1 else ""
    if not source:
        sys.exit("Error: Submitter name not found.")

    column_names = data[5]
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in column_names]
    if missing_cols:
        sys.exit(f"Error: Missing required columns: {', '.join(missing_cols)}")

    column_indices = {col: column_names.index(col) for col in REQUIRED_COLUMNS}
    parsed_data = defaultdict(dict)

    for row_idx, row in enumerate(data[6:], start=7):
        if not row or len(row) <= max(column_indices.values()):
            continue
        try:
            sra_id             = row[column_indices['SRA ID']]
            sp_sg              =  row[column_indices['Assembly Species / Segment']]
            if not sp_sg:
                continue     
            sp_sg_parts        = [ i.lower().strip() for i in sp_sg.split('/') ]
            species            = sp_sg_parts[0]
            segment            = sp_sg_parts[1] if len(sp_sg_parts) > 1 else 'wg'
            assembly_generated = str(row[column_indices['Assembly Generated?']] or '').lower().strip()

            if not sra_id or not species or not segment or assembly_generated not in ASSEMBLY_GENERATED_VALUES:
                continue

            parsed_data[str(sra_id)][(species, segment)] = {
                'name':          row[column_indices['Sample ID']],
                'file':          row[column_indices['Assembly Filename']],
                'source':     source,
                'compute_env':   str(row[column_indices['Compute Environment (HPC cluster, cloud/local workstation)']] or '').lower().strip(),
                'workflow':      str(row[column_indices['Workflow Name & Version']] or '').lower().strip(),
                'workflow_link': str(row[column_indices['Workflow Github Link (if hosted/public on GH)']] or '').lower().strip(),
                'workflow_alt':     'not specified',
                'workflow_version': 'not specified',
            }

        except (IndexError, KeyError) as e:
            logger.warning(f"Skipping row {row_idx} due to parsing error: {e}")

    if not parsed_data:
        sys.exit("Error: No valid assembly data found in workbook.")

    return source, dict(parsed_data)


def load_workflow_map(path: str) -> Dict[str, Tuple[str, str]]:
    """
    Load a CSV mapping workflow names to their canonical name and version.
    Expected columns: workflow, workflow_alt, version
    Returns: {workflow: (workflow_alt, version)}
    """
    workflow_map = {}
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            workflow = (row.get('workflow') or '').lower().strip()
            alt      = (row.get('workflow_alt') or '').lower().strip()
            version  = (row.get('version') or '').lower().strip()
            if workflow:
                workflow_map[workflow] = (alt, version)
    return workflow_map


def apply_workflow_map(data: Dict, workflow_map: Dict) -> Dict:
    for sra_id, sp_sg in data.items():
        for (species, segment), info in sp_sg.items():
            workflow = info.get('workflow', '')
            if not workflow:
                continue
            if workflow not in workflow_map:
                raise ValueError(f"Workflow '{workflow}' (SRA: {sra_id}, species: {species}, segment: {segment}) not found in workflow map.")
            info['workflow_alt'], info['workflow_version'] = workflow_map[workflow]
    return data


def stage_gzip_to_outdir(src: Path, dst_dir: Path) -> Path:
    """Copy src into dst_dir as a gzip file. Compresses if not already .gz."""
    dst_dir.mkdir(parents=True, exist_ok=True)

    if src.suffix == ".gz":
        dst = dst_dir / src.name
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)
        return dst

    dst = dst_dir / (src.name + ".gz")
    with src.open("rb") as f_in, gzip.open(dst, "wb") as f_out:
        shutil.copyfileobj(f_in, f_out)
    return dst


def validate_and_process_fasta_files(data: Dict, fasta_files: List[str], outdir: Path, ignore_missing: bool) -> Dict:
    """Validate that all required FASTA files exist and stage gzipped copies in outdir/fasta."""
    provided_paths = [Path(p).expanduser().resolve() for p in fasta_files]
    for p in provided_paths:
        if not p.exists():
            sys.exit(f"Error: FASTA file not found: {p}")

    by_full = {str(p): p for p in provided_paths}
    by_base: Dict[str, List[Path]] = defaultdict(list)
    for p in provided_paths:
        by_base[p.name].append(p)

    dupes = {b: lst for b, lst in by_base.items() if len(lst) > 1}
    if dupes:
        detail = "\n".join(f"  {b}:\n    " + "\n    ".join(map(str, lst)) for b, lst in dupes.items())
        sys.exit(
            "Error: Duplicate basenames among provided FASTA files (ambiguous workbook references).\n"
            "Use unique filenames or supply full paths in the workbook.\n" + detail
        )

    fasta_outdir = outdir / "fasta"

    for sra_id, sp_sg in data.items():
        for (species, segment), info in sp_sg.items():
            ref = str(info.get('file') or '').strip()
            if not ref:
                logger.warning(f"Empty filename for {sra_id}/{species}_{segment}; skipping.")
                continue

            ref_path = Path(ref)
            candidate = None
            if ref_path.exists():
                candidate = ref_path.resolve()
            elif ref in by_full:
                candidate = by_full[ref]
            elif ref_path.name in by_base:
                candidate = by_base[ref_path.name][0]

            if candidate is None or not candidate.exists():
                if ignore_missing:
                    continue
                sys.exit(f'Error: Required file "{ref}" (for {sra_id}/{segment}) not found among --fasta files.')

            info['file'] = str(stage_gzip_to_outdir(candidate, fasta_outdir))

    return data


def export_data(data: Dict, source: str, outdir: Path) -> None:
    stem = sanitize_string(source)

    # json_path = outdir / f'{stem}.json'
    # try:
    #     with json_path.open('w', encoding='utf-8') as f:
    #         json.dump(data, f, indent=4, ensure_ascii=False)
    #     logger.info(f'Data exported to {json_path}')
    # except Exception as e:
    #     sys.exit(f"Error: Could not write {json_path}: {e}")

    csv_rows = [
        {
            'sample':            f'{sanitize_string(sra_id)}_{sanitize_string(species)}' + ('' if segment == 'wg' else sanitize_string(segment)),
            'species':           sanitize_string(species),
            'segment':           sanitize_string(segment),
            'assembly':          info['file'],
            'source':         source,
            'workflow':          info['workflow'],
            'workflow_alt':      info['workflow_alt'],
            'workflow_version':  info['workflow_version'],
            'compute_env':       info['compute_env'],
        }
        for sra_id, sp_sg in data.items()
        for (species, segment), info in sp_sg.items()
    ]

    csv_path = outdir / f'{stem}.csv'
    try:
        with csv_path.open('w', newline='', encoding='utf-8') as f:
            if csv_rows:
                writer = csv.DictWriter(f, fieldnames=csv_rows[0].keys())
                writer.writeheader()
                writer.writerows(csv_rows)
        logger.info(f'Data exported to {csv_path}')
    except Exception as e:
        sys.exit(f"Error: Could not write {csv_path}: {e}")


def main():
    parser = argparse.ArgumentParser(
        description="Process genomic assembly metadata and FASTA files for submission preparation.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--meta",           required=True,            help="Path to Excel metadata file")
    parser.add_argument("--fasta",          required=True, nargs='+', help="FASTA file paths (space-separated)")
    parser.add_argument("--workflow-map",                             help="CSV mapping workflow names to canonical names/versions")
    parser.add_argument("--outdir",         default=".",              help="Output directory (default: current directory)")
    parser.add_argument("--ignore-missing", action="store_true",      help="Ignore missing assembly files")
    parser.add_argument("--version",        action="version", version=f"prepare-entry.py v{VERSION}")
    args = parser.parse_args()

    outdir = Path(args.outdir).expanduser().resolve()
    ensure_outdir(outdir)

    print(textwrap.dedent(f"""
        prepare-entry.py v{VERSION}
        Output directory: {outdir}
        Processing genomic assembly data...
    """), flush=True)

    if not os.path.exists(args.meta):
        sys.exit(f"Error: Metadata file not found: {args.meta}")

    logger.info("Loading workbook data...")
    source, data = load_workbook_data(args.meta)
    logger.info(f"source: {source} | SRA entries: {len(data)}")

    if args.workflow_map:
        logger.info("Applying workflow map...")
        workflow_map = load_workflow_map(args.workflow_map)
        try:
            data = apply_workflow_map(data, workflow_map)
        except ValueError as e:
            sys.exit(f"Error: {e}")

    logger.info("Validating and staging FASTA files...")
    data = validate_and_process_fasta_files(data, args.fasta, outdir, args.ignore_missing)

    logger.info("Exporting processed data...")
    export_data(data, source, outdir)

    logger.info("Done.")


if __name__ == "__main__":
    main()